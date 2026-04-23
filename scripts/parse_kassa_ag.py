"""
parse_kassa_ag.py — читает УЧЕТ НАЛИЧНЫХ АГ xlsx (несколько помесячных листов)
и пишет data/snapshots/ag_kassa.json со списком {month, in, out, ending}.

Структура каждого листа (ДЕКАБРЬ 2025 / ЯНВАРЬ / ФЕВРАЛЬ / МАРТ / АПРЕЛЬ):
  R6:     "Переходящий остаток" (col D = начальный остаток EUR)
  R~23:   "Всего:" в левой части (col D = приход), в правой (col F + col G = расход)
  R~24:   "ИТОГО:" (col G = итог расход)
  R~26:   "Переходящий остаток на ..." (col F = остаток на след. месяц)

Usage:  python scripts/parse_kassa_ag.py _raw/УЧЕТ\ НАЛИЧНЫХ\ 2026\ АГ.xlsx
"""
import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

MONTH_MAP = {
    "ДЕКАБРЬ 2025": "2025-12",
    "ЯНВАРЬ": "2026-01",
    "ФЕВРАЛЬ": "2026-02",
    "МАРТ": "2026-03",
    "АПРЕЛЬ": "2026-04",
    "МАЙ": "2026-05", "ИЮНЬ": "2026-06", "ИЮЛЬ": "2026-07",
    "АВГУСТ": "2026-08", "СЕНТЯБРЬ": "2026-09", "ОКТЯБРЬ": "2026-10",
    "НОЯБРЬ": "2026-11", "ДЕКАБРЬ 2026": "2026-12",
}


def num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


def parse_sheet(ws) -> dict:
    """Структура меняется месяц к месяцу. Универсально: найти строки с 'Всего:',
    собрать все числа в той же строке. Левый столбец → in, правый → out."""
    rows = list(ws.iter_rows(values_only=True))
    result = {"start": None, "in": None, "out_total": None, "ending": None}

    for row in rows:
        row = list(row)
        # Начальный остаток: "Переходящий остаток за ..." → ищем число в строке
        text_cells = [(i, str(v)) for i, v in enumerate(row) if v is not None]
        full_text = " | ".join(s for _, s in text_cells)

        if "Переходящий остаток за" in full_text and result["start"] is None:
            # число обычно в col D (index 3)
            for v in row:
                n = num(v)
                if n and n > 0:
                    result["start"] = n
                    break

        # Переходящий остаток на ... (ending)
        if "Переходящий остаток на" in full_text:
            for v in row:
                n = num(v)
                if n is not None:
                    result["ending"] = n
                    break

        # "Всего:" — найти все численные значения в строке
        if any("Всего:" in s or "ИТОГО" in s for _, s in text_cells):
            nums = [num(v) for v in row]
            nums = [n for n in nums if n is not None and n > 0]
            vsego_positions = [i for i, s in text_cells if "Всего:" in s or "ИТОГО" in s]
            # Левый "Всего:" (позиция ≤ 4) → result["in"]; правый (позиция ≥ 5) → out
            for col_i in vsego_positions:
                # Сканируем от "Всего:" до ближайшей пустой ячейки (разделителя).
                # Левый: forward (col D единственная, потом col E пусто → stop)
                # Правый: backward (col F, иногда + col G, потом col E пусто → stop)
                total = 0.0
                if col_i <= 4:  # left side — forward
                    for v_i in range(col_i + 1, len(row)):
                        if row[v_i] is None or row[v_i] == "":
                            break
                        n = num(row[v_i])
                        if n and n > 0:
                            total += n
                else:  # right side — backward
                    for v_i in range(col_i - 1, -1, -1):
                        if row[v_i] is None or row[v_i] == "":
                            break
                        n = num(row[v_i])
                        if n and n > 0:
                            total += n
                if total == 0: continue
                if col_i <= 4 and result["in"] is None:
                    result["in"] = total
                elif col_i >= 5 and result["out_total"] is None:
                    result["out_total"] = total

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_kassa_ag.py <xlsx_path>")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    months = []
    for sheet_name in wb.sheetnames:
        period = MONTH_MAP.get(sheet_name.strip().upper())
        if not period:
            print(f"skip sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        parsed = parse_sheet(ws)
        parsed["month"] = period
        parsed["sheet"] = sheet_name
        months.append(parsed)

    wb.close()
    months.sort(key=lambda x: x["month"])

    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(root, "data", "snapshots", "ag_kassa.json")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({"months": months, "source_file": os.path.basename(xlsx_path)},
                  f, ensure_ascii=False, indent=2)
    print(f"Wrote {out_path}")
    for m in months:
        fi = f"{m['in']:.0f}" if m['in'] else "-"
        fo = f"{m['out_total']:.0f}" if m['out_total'] else "-"
        fe = f"{m['ending']:.0f}" if m['ending'] else "-"
        print(f"  {m['month']}: in={fi}  out={fo}  end={fe}")


if __name__ == "__main__":
    main()
